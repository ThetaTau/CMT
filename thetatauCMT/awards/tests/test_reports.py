import csv
import datetime
import io

import pytest
from django.contrib.auth.models import Group
from django.urls import reverse

from thetatauCMT.awards import reports
from thetatauCMT.awards.services import revoke_grant
from thetatauCMT.awards.tests._helpers import sign_rmp as _sign_rmp
from thetatauCMT.awards.tests.factories import AwardCycleFactory, AwardGrantFactory, AwardTypeFactory
from thetatauCMT.chapters.models import GREEK_ABR
from thetatauCMT.chapters.tests.factories import ChapterFactory
from thetatauCMT.regions.tests.factories import RegionFactory
from thetatauCMT.users.tests.factories import UserFactory

pytestmark = pytest.mark.django_db

_NAMES = list(GREEK_ABR.values())


# ---------------------------------------------------------------------------
# Officer / RMP helpers (mirrors the direct-grant view tests)
# ---------------------------------------------------------------------------
def _officer():
    user = UserFactory()
    user.groups.add(Group.objects.get_or_create(name="officer")[0])
    _sign_rmp(user)
    return user


def _natoff():
    user = UserFactory()
    user.groups.add(Group.objects.get_or_create(name="natoff")[0])
    _sign_rmp(user)
    return user


def _non_officer():
    user = UserFactory()
    _sign_rmp(user)
    return user


def _csv_rows(response):
    return list(csv.reader(response.content.decode().splitlines()))


# ===========================================================================
# Report queries
# ===========================================================================
def test_awards_by_cycle_active_only_by_default():
    cycle = AwardCycleFactory()
    keep = AwardGrantFactory(cycle=cycle)
    other = AwardGrantFactory(cycle=AwardCycleFactory())
    revoked = AwardGrantFactory(cycle=cycle)
    revoke_grant(revoked, revoked_by=UserFactory(), reason="x")
    result = list(reports.awards_by_cycle(cycle))
    assert keep in result
    assert other not in result
    assert revoked not in result  # active-only default
    assert revoked in list(reports.awards_by_cycle(cycle, include_revoked=True))


def test_awards_by_award_type():
    award = AwardTypeFactory()
    keep = AwardGrantFactory(award_type=award)
    other = AwardGrantFactory(award_type=AwardTypeFactory())
    result = list(reports.awards_by_award_type(award))
    assert keep in result
    assert other not in result


def test_awards_by_chapter_matches_chapter_and_members():
    chapter = ChapterFactory(name=_NAMES[0])
    other_chapter = ChapterFactory(name=_NAMES[1])
    chapter_grant = AwardGrantFactory(recipient_member=None, recipient_chapter=chapter)
    member_grant = AwardGrantFactory(recipient_member=UserFactory(status="active", chapter=chapter))
    outside = AwardGrantFactory(recipient_member=None, recipient_chapter=other_chapter)
    result = list(reports.awards_by_chapter(chapter))
    assert chapter_grant in result
    assert member_grant in result
    assert outside not in result


def test_awards_by_region_matches_region_chapter_and_member():
    region = RegionFactory(name="Reg One")
    other_region = RegionFactory(name="Reg Two")
    chapter = ChapterFactory(name=_NAMES[0], region=region)
    other_chapter = ChapterFactory(name=_NAMES[1], region=other_region)
    region_grant = AwardGrantFactory(recipient_member=None, recipient_region=region)
    chapter_grant = AwardGrantFactory(recipient_member=None, recipient_chapter=chapter)
    member_grant = AwardGrantFactory(recipient_member=UserFactory(status="active", chapter=chapter))
    outside = AwardGrantFactory(recipient_member=None, recipient_chapter=other_chapter)
    result = list(reports.awards_by_region(region))
    assert region_grant in result
    assert chapter_grant in result
    assert member_grant in result
    assert outside not in result


def test_member_award_history_ordered_by_effective_date_incl_backdated_and_revoked():
    member = UserFactory(status="active")
    recent = AwardGrantFactory(recipient_member=member, effective_date=datetime.date(2025, 1, 1))
    backdated = AwardGrantFactory(recipient_member=member, effective_date=datetime.date(2010, 1, 1))
    middle = AwardGrantFactory(recipient_member=member, effective_date=datetime.date(2018, 6, 1))
    revoke_grant(recent, revoked_by=UserFactory(), reason="x")
    history = list(reports.member_award_history(member))
    # Chronological by effective_date; backdated first.
    assert history == [backdated, middle, recent]
    # Revoked grant is still part of the full history.
    assert recent.is_revoked


def test_chapter_award_history_ordered():
    chapter = ChapterFactory(name=_NAMES[0])
    old = AwardGrantFactory(recipient_member=None, recipient_chapter=chapter, effective_date=datetime.date(2015, 1, 1))
    new = AwardGrantFactory(
        recipient_member=UserFactory(status="active", chapter=chapter), effective_date=datetime.date(2024, 1, 1)
    )
    assert list(reports.chapter_award_history(chapter)) == [old, new]


# ===========================================================================
# Exports: correct rows (CSV + XLSX)
# ===========================================================================
def test_csv_export_returns_correct_rows(client):
    cycle = AwardCycleFactory()
    a = AwardGrantFactory(award_type=AwardTypeFactory(name="Row Award A"), cycle=cycle)
    b = AwardGrantFactory(award_type=AwardTypeFactory(name="Row Award B"), cycle=cycle)
    AwardGrantFactory(award_type=AwardTypeFactory(name="Other Cycle Award"))  # different cycle
    client.force_login(_officer())
    response = client.get(reverse("awards:export"), {"cycle": cycle.pk})
    assert response.status_code == 200
    assert "text/csv" in response["Content-Type"]
    rows = _csv_rows(response)
    assert rows[0][0] == "Award"  # header
    awards_in_csv = {row[0] for row in rows[1:]}
    assert awards_in_csv == {"Row Award A", "Row Award B"}
    recipients = {row[3] for row in rows[1:]}
    assert a.recipient_display in recipients
    assert b.recipient_display in recipients


def test_csv_export_excludes_revoked_by_default_and_includes_on_request(client):
    cycle = AwardCycleFactory()
    AwardGrantFactory(award_type=AwardTypeFactory(name="Active Row Award"), cycle=cycle)
    revoked = AwardGrantFactory(award_type=AwardTypeFactory(name="Revoked Row Award"), cycle=cycle)
    revoke_grant(revoked, revoked_by=UserFactory(), reason="x")
    client.force_login(_officer())
    default_rows = _csv_rows(client.get(reverse("awards:export"), {"cycle": cycle.pk}))
    assert {r[0] for r in default_rows[1:]} == {"Active Row Award"}
    with_revoked = _csv_rows(client.get(reverse("awards:export"), {"cycle": cycle.pk, "include_revoked": "1"}))
    assert {r[0] for r in with_revoked[1:]} == {"Active Row Award", "Revoked Row Award"}


def test_xlsx_export_returns_workbook_with_rows(client):
    from openpyxl import load_workbook

    cycle = AwardCycleFactory()
    AwardGrantFactory(award_type=AwardTypeFactory(name="Excel Award"), cycle=cycle)
    client.force_login(_officer())
    response = client.get(reverse("awards:export"), {"cycle": cycle.pk, "format": "xlsx"})
    assert response.status_code == 200
    assert "spreadsheetml" in response["Content-Type"]
    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    values = [[cell.value for cell in row] for row in sheet.iter_rows()]
    assert values[0][0] == "Award"
    assert any(row[0] == "Excel Award" for row in values[1:])


def test_export_by_chapter_and_member(client):
    chapter = ChapterFactory(name=_NAMES[0])
    member = UserFactory(status="active", chapter=chapter)
    AwardGrantFactory(award_type=AwardTypeFactory(name="Chapter Member Award"), recipient_member=member)
    AwardGrantFactory(award_type=AwardTypeFactory(name="Elsewhere Award"))
    client.force_login(_officer())
    chapter_rows = _csv_rows(client.get(reverse("awards:export"), {"chapter": chapter.slug}))
    assert {r[0] for r in chapter_rows[1:]} == {"Chapter Member Award"}
    member_rows = _csv_rows(client.get(reverse("awards:export"), {"member": member.username}))
    assert {r[0] for r in member_rows[1:]} == {"Chapter Member Award"}


# ===========================================================================
# Exports: permission gating
# ===========================================================================
def test_export_redirects_anonymous(client):
    assert client.get(reverse("awards:export")).status_code == 302


def test_export_blocks_non_officer(client):
    client.force_login(_non_officer())
    response = client.get(reverse("awards:export"))
    assert response.status_code == 302
    assert "text/csv" not in response.get("Content-Type", "")


def test_export_allows_officer(client):
    AwardGrantFactory()
    client.force_login(_officer())
    response = client.get(reverse("awards:export"))
    assert response.status_code == 200
    assert "text/csv" in response["Content-Type"]


# ===========================================================================
# History views (public)
# ===========================================================================
def test_member_history_view_public_and_ordered(client):
    member = UserFactory(status="active")
    AwardGrantFactory(
        award_type=AwardTypeFactory(name="History Award"),
        recipient_member=member,
        effective_date=datetime.date(2012, 1, 1),
    )
    response = client.get(reverse("awards:member_history", kwargs={"username": member.username}))
    assert response.status_code == 200  # public, no login
    assert "Award history" in response.content.decode()
    history = list(response.context["object_list"])
    assert history == list(reports.member_award_history(member))


def test_member_history_includes_revoked_for_national_officer(client):
    member = UserFactory(status="active")
    grant = AwardGrantFactory(award_type=AwardTypeFactory(name="Revoked History Award"), recipient_member=member)
    revoke_grant(grant, revoked_by=UserFactory(), reason="x")
    client.force_login(_natoff())
    response = client.get(reverse("awards:member_history", kwargs={"username": member.username}))
    assert grant in list(response.context["object_list"])
    assert "Revoked" in response.content.decode()  # status column labels it


def test_member_history_hides_revoked_from_non_national_officer(client):
    member = UserFactory(status="active")
    grant = AwardGrantFactory(award_type=AwardTypeFactory(name="Hidden Revoked History"), recipient_member=member)
    revoke_grant(grant, revoked_by=UserFactory(), reason="x")
    response = client.get(reverse("awards:member_history", kwargs={"username": member.username}))
    assert grant not in list(response.context["object_list"])


def test_chapter_history_view_public(client):
    chapter = ChapterFactory(name=_NAMES[0])
    AwardGrantFactory(
        award_type=AwardTypeFactory(name="Chapter History Award"), recipient_member=None, recipient_chapter=chapter
    )
    response = client.get(reverse("awards:chapter_history", kwargs={"slug": chapter.slug}))
    assert response.status_code == 200
    assert "Chapter History Award" in response.content.decode()


def test_history_export_buttons_officer_only(client):
    member = UserFactory(status="active")
    AwardGrantFactory(recipient_member=member)
    url = reverse("awards:member_history", kwargs={"username": member.username})
    # Anonymous: no export controls.
    assert client.get(url).context["can_export"] is False
    # Officer: export controls available.
    client.force_login(_officer())
    officer_response = client.get(url)
    assert officer_response.context["can_export"] is True
    # CSV export is available; the Excel export button was removed.
    assert reverse("awards:export") in officer_response.content.decode()
    assert "format=xlsx" not in officer_response.content.decode()
